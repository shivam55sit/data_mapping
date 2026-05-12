import streamlit as st
import pandas as pd
import os
import io
from decode_col import decode
from datetime import datetime

# ── PAGE CONFIG ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Excel Standardiser",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── CUSTOM CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    /* Global */
    .stApp {
        font-family: 'Inter', sans-serif;
    }

    /* Hero header */
    .hero {
        background: linear-gradient(135deg, #0f0c29 0%, #302b63 50%, #24243e 100%);
        border-radius: 16px;
        padding: 2.5rem 2rem;
        margin-bottom: 2rem;
        text-align: center;
        box-shadow: 0 8px 32px rgba(48, 43, 99, 0.3);
    }
    .hero h1 {
        color: #ffffff;
        font-size: 2.2rem;
        font-weight: 700;
        margin: 0 0 0.5rem 0;
        letter-spacing: -0.5px;
    }
    .hero p {
        color: #b8b5d4;
        font-size: 1.05rem;
        margin: 0;
        font-weight: 300;
    }

    /* Stats cards */
    .stat-row {
        display: flex;
        gap: 1rem;
        margin: 1.5rem 0;
    }
    .stat-card {
        flex: 1;
        background: linear-gradient(145deg, #1a1a2e, #16213e);
        border: 1px solid rgba(99, 102, 241, 0.2);
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        text-align: center;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    .stat-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(99, 102, 241, 0.15);
    }
    .stat-number {
        font-size: 2rem;
        font-weight: 700;
        background: linear-gradient(135deg, #667eea, #764ba2);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin: 0;
    }
    .stat-label {
        color: #8b8fa3;
        font-size: 0.85rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        margin: 0.3rem 0 0 0;
    }

    /* Upload area */
    .upload-zone {
        border: 2px dashed rgba(99, 102, 241, 0.4);
        border-radius: 16px;
        padding: 2rem;
        text-align: center;
        background: rgba(99, 102, 241, 0.03);
        transition: border-color 0.3s;
        margin-bottom: 1.5rem;
    }
    .upload-zone:hover {
        border-color: rgba(99, 102, 241, 0.7);
    }

    /* Log entries */
    .log-entry {
        padding: 0.5rem 1rem;
        margin: 0.3rem 0;
        border-radius: 8px;
        font-size: 0.9rem;
        font-family: 'SF Mono', 'Fira Code', monospace;
    }
    .log-success {
        background: rgba(16, 185, 129, 0.1);
        border-left: 3px solid #10b981;
        color: #10b981;
    }
    .log-info {
        background: rgba(99, 102, 241, 0.1);
        border-left: 3px solid #6366f1;
        color: #818cf8;
    }
    .log-warn {
        background: rgba(245, 158, 11, 0.1);
        border-left: 3px solid #f59e0b;
        color: #fbbf24;
    }

    /* Section headers */
    .section-header {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        margin: 1.5rem 0 1rem 0;
    }
    .section-header h3 {
        margin: 0;
        font-weight: 600;
        font-size: 1.15rem;
    }

    /* Badges */
    .badge {
        display: inline-block;
        padding: 0.2rem 0.7rem;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .badge-purple {
        background: rgba(99, 102, 241, 0.15);
        color: #818cf8;
    }
    .badge-green {
        background: rgba(16, 185, 129, 0.15);
        color: #10b981;
    }
    .badge-orange {
        background: rgba(245, 158, 11, 0.15);
        color: #f59e0b;
    }

    /* Mapping table */
    .mapping-table {
        width: 100%;
        border-collapse: separate;
        border-spacing: 0;
        border-radius: 12px;
        overflow: hidden;
        margin: 1rem 0;
    }
    .mapping-table th {
        background: rgba(99, 102, 241, 0.12);
        padding: 0.75rem 1rem;
        text-align: left;
        font-weight: 600;
        font-size: 0.85rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        color: #818cf8;
    }
    .mapping-table td {
        padding: 0.6rem 1rem;
        font-size: 0.9rem;
        border-bottom: 1px solid rgba(255,255,255,0.05);
    }
    .mapping-table tr:last-child td {
        border-bottom: none;
    }
    .arrow-cell {
        text-align: center;
        color: #6366f1;
        font-weight: 700;
        font-size: 1.1rem;
    }

    /* Hide default streamlit elements */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    .stDeployButton {display: none;}
    [data-testid="stSidebar"] {display: none;}

    /* Download button style */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        border: none !important;
        padding: 0.75rem 2rem !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        transition: transform 0.2s, box-shadow 0.2s !important;
        width: 100% !important;
    }
    .stDownloadButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 25px rgba(102, 126, 234, 0.4) !important;
    }
</style>
""", unsafe_allow_html=True)


# ── REFERENCE FILE ──────────────────────────────────────────────────────────
REFERENCE_FILE = "Demonstration Multipurpose School.xls"

# ── COLUMN MAPS ─────────────────────────────────────────────────────────────
TRIAGE_MAP = {
    'School Name ': 'School Name',
    'School code:': 'School Code',
    'Class ': 'Class',
    'Section': 'Section',
    'Subject ID': 'Child Unique Code',
    'Id Number' : 'Child Unique Code',
    'Parent Consent (Yes/No)': 'Parent Consent',
    'Name of the child': 'Student Name',
    'Father/Guardian Name': 'Parent Or Guardians Name',
    'Mobile number': 'Contact Number',
    'Available (1=Yes;0=No)': 'Available',
    'Age (Completed Years)': 'Age',
    'Gender(0: Male, 1: Female)': 'Gender',
    'Any disability*': 'Disability Any',
    'Using glasses?': 'Using Glasses',
    'Vision test Right Eye': 'Vision Test Right Eye',
    'Vision test Left Eye': 'Vision Test Left Eye',
    'Signs & Symptoms': 'SS Normal',
    'Referral*(0=No; 1=Yes)': 'Auto Referral',
    'Reason for Referral,' : 'Reason For Referral'
}




VCH_MAP = {
    "ID Number": "Child Unique Code",
    "Name of the child": "Student Name",
    "Name of the Father/Guardian": "Parent Or Guardians Name",
    "Gender  Male:0, Female:1": "Gender",
    "Age": "Age",
    "Date": "Date Of Examination",
    "Name of the VT/Optom": "Examiner Name",
    "Additional information": "Remarks",
    "OD Unaided Distance": "Va OD Unaided",
    "OD Pinhole Distance": "Va OD Pin Hole",
    "OS Unaided Distance": "Va OS Unaided",
    "OS Pinhole Distance": "Va OS Pin Hole",
    "OD Aided Distance": "Va OD Aided Glasses",
    "OS Aided Distance": "Va OS Aided Glasses",
    "Present glasses (0=No, 1=Yes)": "Using Glasses",
    "OD Spherical  Objective refraction": "Sph Or RE",
    "OD Cylindrical Objective refraction": "Cyl Or RE",
    "OD Axis Objective refraction": "Axis Or RE",
    "OS Spherical Objective refraction": "Sph Or LE",
    "OS Cylindrical Objective refraction": "Cyl Or LE",
    "OS Axis Objective refraction": "Axis Or LE",
    "OD Spherical Acceptance": "Sph Acceptance RE",
    "OD Cylindircal Acceptance": "Cyl Acceptance RE",
    "OD Axis Acceptance": "Axis Acceptance RE",
    "OD VA Acceptance": "C2 Va Accptance RE",
    "OS Spherical Acceptance": "Sph Acceptance LE",
    "OS Cylindircal  Acceptance": "Cyl Acceptance LE",
    "OS Axis  Acceptance": "Axis Acceptance LE",
    "OS VA  Acceptance": "C2 Va Accptance LE",
    "OD Spherical Lensometry": "Sph Lensometry RE",
    "OD Cylindircal Lensometry": "Cyl Lensometry RE",
    "OD Axis Lensometry": "Axis Lensometry RE",
    "OS Spherical Lensometry": "Sph Lensometry LE",
    "OS Cylindircal Lensometry": "Cyl Lensometry LE",
    "OS Axis Lensometry": "Axis Lensometry LE",
    "Lids OD": "Slit Lids OD",
    "Specify.2": "C6 Lids Others RE",
    "Lids OS": "Slit Lids OS",
    "Specify.3": "C6 Lids Others LE",
    "Conjunctiva OD": "Slit Conjunctiva OD C6",
    "Specify.4": "C6 Conjunctiva Others RE",
    "Conjunctiva OS": "Slit Conjunctiva OS C6",
    "Specify.5": "C6 Conjunctiva Others LE",
    "Cornea OD": "Slit Cornea OD C6",
    "Specify.6": "C6 Cornea Others RE",
    "Cornea OS": "Slit Cornea OS C6",
    "Specify.7": "C6 Cornea Others LE",
    "Anterior Chamber OD": "Slit AC Status OD C6",
    "Specify.8": "Slit AC Status OD C6 Other",
    "Anterior Chamber OS": "Slit AC Status OS C6",
    "Specify.9": "Slit AC Status OS C6 Other",
    "Iris/Pupil OD": "Slit Pupil Status OD C6",
    "Specify.10": "Slit Pupil Status OD C6 Other",
    "Iris/Pupil OS": "Slit Pupil Status OS C6",
    "Specify.11": "Slit Pupil Status OS C6 Other",
    "lens OD": "Slit Lens Status OD C6",
    "Specify.12": "C6 Lens RE Others",
    "lens OS": "Slit Lens Status OS C6",
    "Specify.13": "C6 Lens LE Others",
    "Redd reflex OD": "Red Reflex OD C7",
    "Specify.14": "C7 Red Reflex Others RE",
    "Redd reflex OS": "Red Reflex OS C7",
    "Specify.15": "C7 Red Reflex Others LE",
    "Diagnosis OD": "Right Impression",
    "Diagnosis OS": "Left Impression",
    "Cause of Visual impairement OD": "Right Refractive Error",
    "Cause of Visual impairement OS": "Left Refractive Error",
    # Cover Test / Squint / Nystagmus
    "Cover Test": "SS Squint",
    "Nystagmus": "Shaking Eye Ball",
    # Fundus
    "Fundus OD": "Right Retinal Evaluation",
    "Fundus OS": "Left Retinal Evaluation",
    # Topography
    "Topography OD": "Right Asymmetrical Topography",
    "Topography OS": "Left Asymmetrical Topography",
    # Action / Referral
    "Action": "Advise For Next Module",
    "Name of the TC/ City centre/VC": "C11 Vision Center Name",
    # Date of last eye check
    "Date of last eye Check up": "Last Eye Check B1",
}

# GHS_MAP keys are PREFIX strings — actual Excel headers may be longer.
GHS_MAP = {
    "Child Id": "Child Unique Code",
    "Name of the child": "Student Name",
    "Name of the Father/Guardian": "Parent Or Guardians Name",
    "Date of examination": "Clinical Examination Date",
    "1.When was the last time you got your eyes checked?": "Last Eye Check B1",
    "2. Do you have the habit of rubbbing eyes?": "Rub Eyes B2",
    "3.How frequency do you rub your eyes?": "Rub Freq B3",
    "4. Why do you rub your eyes?": "Rub Reason B4",
    "5.when you rub your eyes, which of the following do you usually use?": "Rub Method B5",
    "6. Do you think eye rubbing can damage the eyes?": "Think Damage Rub B6",
    "8. Do you frequently encounter redness in your eye?": "Eye Redness B8",
    "9. If Yes, how frequently?": "Redness Frequency B9",
    "10.Do you feel any itching/scratching sensation in your eyes?": "Eye Itching B10",
    "11.If Yes, How frequently?": "Itching Frequency B11",
    "12. Do you ever have difficulty completing school/work": "Difficulty School Work B12",
    "13. Do you have trouble paying attention": "Attention School B13",
    "14.Do you miss school": "School Absent B14",
    "15.Do you have the habit of sleeping with your face down?": "Sleep Face Down B15",
    "16. Do you have episodes of sneezing": "Sneezing Breath B16",
    "17.How frequently?": "Sneeze Freq B17",
    "18. Do you have any history of recurrent itching": "Skin Itch Rash B18",
    "19.Are you on medications": "Medication Allergy B19",
    "20. If yes, do you recall what medications you are taking": "Medication Name B20",
    "21. Are you left/right handed?": "Handedness B21",
    "22.Haveyou undergone any eye surgeries before?": "Eye Surgery B22",
}


# ── HELPER FUNCTIONS ────────────────────────────────────────────────────────
@st.cache_data
def get_standard_columns():
    """Read standard column order from the tablet-recorded reference file."""
    ref_df = pd.read_excel(REFERENCE_FILE, nrows=0)
    return list(ref_df.columns)


def normalize_multiline_columns(df):
    """Extract only the first line of multi-line column headers."""
    df.columns = [str(col).split("\n")[0].strip() for col in df.columns]
    return df


def rename_by_prefix(df, col_map):
    """
    Rename columns using prefix matching.
    For each map key, if a column starts with that key text, rename it.
    Skips pandas dedup suffixes like '.1', '.2' to avoid false matches.
    """
    rename_dict = {}
    sorted_keys = sorted(col_map.keys(), key=len, reverse=True)
    for actual_col in df.columns:
        for map_key in sorted_keys:
            if actual_col == map_key:
                rename_dict[actual_col] = col_map[map_key]
                break
            remaining = actual_col[len(map_key):]
            if actual_col.startswith(map_key) and remaining and not remaining[0].isalnum() and not remaining.startswith('.'):
                rename_dict[actual_col] = col_map[map_key]
                break
    df = df.rename(columns=rename_dict)
    return df


def process_uploaded_file(uploaded_file, standard_columns):
    """Process a single uploaded manual Excel file and return standardised DataFrame + logs."""
    logs = []

    # Read both sheets
    triage = pd.read_excel(uploaded_file, sheet_name=0, header=1)
    vch = pd.read_excel(uploaded_file, sheet_name=1, header=1)
    ghs = pd.read_excel(uploaded_file, sheet_name=2, header=1)

    triage = normalize_multiline_columns(triage)
    vch = normalize_multiline_columns(vch)
    ghs = normalize_multiline_columns(ghs)

    logs.append(("info", f"Triage sheet: {triage.shape[0]} rows × {triage.shape[1]} columns"))
    logs.append(("info", f"VCH sheet: {vch.shape[0]} rows × {vch.shape[1]} columns"))
    logs.append(("info", f"GHS sheet: {ghs.shape[0]} rows × {ghs.shape[1]} columns"))

    # Rename: exact match for VCH, prefix match for GHS (GHS has longer headers)
    triage = triage.rename(columns=TRIAGE_MAP)
    vch = vch.rename(columns=VCH_MAP)
    ghs = rename_by_prefix(ghs, GHS_MAP)

    triage_mapped = len([c for c in triage.columns if c in set(TRIAGE_MAP.values())])
    vch_mapped = len([c for c in vch.columns if c in set(VCH_MAP.values())])
    ghs_mapped = len([c for c in ghs.columns if c in set(GHS_MAP.values())])

    logs.append(("success", f"Triage: {triage_mapped} columns mapped to standard format"))
    logs.append(("success", f"VCH: {vch_mapped} columns mapped to standard format"))
    logs.append(("success", f"GHS: {ghs_mapped} columns mapped to standard format"))

    # Convert Child Unique Code to string to avoid type mismatch during merge
    if "Child Unique Code" in vch.columns:
        vch["Child Unique Code"] = vch["Child Unique Code"].apply(
            lambda x: str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else str(x) if pd.notna(x) else ""
        )
    if "Child Unique Code" in ghs.columns:
        ghs["Child Unique Code"] = ghs["Child Unique Code"].apply(
            lambda x: str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else str(x) if pd.notna(x) else ""
        )
    if "Child Unique Code" in triage.columns:
            triage["Child Unique Code"] = triage["Child Unique Code"].apply(
                lambda x: str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else str(x) if pd.notna(x) else ""
            )  

    # Merge on Child Unique Code
    if "Child Unique Code" in vch.columns and "Child Unique Code" in ghs.columns and "Child Unique Code" in triage.columns:
        # 1. Merge Triage and VCH
        vch_only_cols = [c for c in vch.columns if c not in triage.columns or c == "Child Unique Code"]
        vch_for_merge = vch[vch_only_cols]
        merged = pd.merge(triage, vch_for_merge, on="Child Unique Code", how="outer")

        # 2. Merge result with GHS
        ghs_only_cols = [c for c in ghs.columns if c not in merged.columns or c == "Child Unique Code"]
        ghs_for_merge = ghs[ghs_only_cols]
        merged = pd.merge(merged, ghs_for_merge, on="Child Unique Code", how="outer")
        
        logs.append(("success", f"Merged Triage + VCH + GHS on Child Unique Code → {merged.shape[0]} rows"))
    else:
        merged = vch
        logs.append(("warn", "Could not merge — 'Child Unique Code' missing. Using VCH only."))

    # Drop duplicate columns from merge (keep first)
    merged = merged.loc[:, ~merged.columns.duplicated(keep='first')]

    # Build final output
    final_df = pd.DataFrame()
    mapped_count = 0
    blank_count = 0

    for col in standard_columns:
        if col in merged.columns:
            final_df[col] = merged[col].values
            mapped_count += 1
        else:
            final_df[col] = ""
            blank_count += 1

    final_df["Reg Id"] = range(1, len(final_df) + 1)

    # Sanitize dataframe to prevent PyArrow mixed-type serialization errors
    # and ensure openpyxl writes clean files
    for c in final_df.columns:
        if final_df[c].dtype == object:
            # Convert non-null values to string, keep empty strings empty
            final_df[c] = final_df[c].apply(
                lambda x: str(x) if pd.notna(x) and str(x).strip() != "" else ""
            )

    logs.append(("success", f"Output: {len(final_df)} rows × {len(standard_columns)} columns"))
    logs.append(("info", f"Columns with data: {mapped_count} | Blank: {blank_count}"))

    return final_df, logs, mapped_count, blank_count


def to_excel_bytes(df):
    """Convert DataFrame to downloadable Excel bytes."""
    from openpyxl.utils import get_column_letter

    # Force long numeric IDs to string so Excel doesn't use scientific notation
    df = df.copy()
    if "Child Unique Code" in df.columns:
        df["Child Unique Code"] = df["Child Unique Code"].apply(
            lambda x: str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else str(x) if pd.notna(x) else ""
        )

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Standardised")

        # Set the "Child Unique Code" column format to Text in Excel
        ws = writer.sheets["Standardised"]
        if "Child Unique Code" in df.columns:
            col_idx = df.columns.get_loc("Child Unique Code") + 1  # +1 for 1-indexed
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):  # skip header
                ws[f"{col_letter}{row}"].number_format = '@'

    return buffer.getvalue()


def decode_dataframe(df):
    """
    Replace encoded numeric values with text labels in the standardised DataFrame.
    Returns (decoded_df, decode_logs, decoded_count).
    """
    df = df.copy()
    decode_logs = []
    decoded_count = 0

    for col_name, value_map in decode.items():
        if col_name in df.columns:
            # Convert column to object dtype so we can safely replace
            df[col_name] = df[col_name].astype(object)
            
            # Build a robust mapping that handles ints, strings, and float-strings ("1.0")
            robust_map = {}
            for k, v in value_map.items():
                robust_map[k] = v
                robust_map[str(k)] = v
                if isinstance(k, (int, float)):
                    robust_map[str(float(k))] = v
            
            # Use pandas replace to swap values based on the robust mapping
            df[col_name] = df[col_name].replace(robust_map)
            decoded_count += 1
            decode_logs.append(("success", f"Decoded: {col_name}"))
        else:
            decode_logs.append(("warn", f"Column not found in data: {col_name}"))

    return df, decode_logs, decoded_count


def find_id_column(sheet_df, col_map):
    """Find the column in the original sheet that maps to 'Child Unique Code'."""
    for orig_name, std_name in col_map.items():
        if std_name == 'Child Unique Code' and orig_name in sheet_df.columns:
            return orig_name
    return None


def fill_missing_from_original(download_df, orig_sheet, id_col_name, ref_col_name, target_col_name):
    """
    Fill blank values in download_df[target_col_name] using orig_sheet[ref_col_name],
    matched on orig_sheet[id_col_name] == download_df['Child Unique Code'].
    Returns (updated_df, filled_count).
    """
    orig_sheet = orig_sheet.copy()
    download_df = download_df.copy()

    normalize_id = lambda x: str(int(x)) if pd.notna(x) and isinstance(x, (int, float)) else str(x).strip() if pd.notna(x) else ""
    orig_sheet[id_col_name] = orig_sheet[id_col_name].apply(normalize_id)
    download_df['Child Unique Code'] = download_df['Child Unique Code'].apply(normalize_id)

    lookup = {}
    for _, row in orig_sheet.iterrows():
        code = row[id_col_name]
        value = row[ref_col_name]
        if code and pd.notna(value) and str(value).strip() != "":
            lookup[code] = value

    if target_col_name not in download_df.columns:
        download_df[target_col_name] = pd.NA

    filled_count = 0
    for idx, row in download_df.iterrows():
        cell = row[target_col_name]
        if pd.isna(cell) or str(cell).strip() == "":
            code = row['Child Unique Code']
            if code in lookup:
                download_df.at[idx, target_col_name] = lookup[code]
                filled_count += 1

    return download_df, filled_count



# ── MAIN CONTENT ────────────────────────────────────────────────────────────

# Hero header
st.markdown("""
<div class="hero">
    <h1>📊 Excel Standardiser</h1>
    <p>Convert manually-recorded school data to the standard tablet format</p>
</div>
""", unsafe_allow_html=True)

# Upload zone
st.markdown('<div class="upload-zone">', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Upload a manually-recorded Excel file (.xlsx or .xls)",
    type=["xlsx", "xls"],
    help="The file should have 2 sheets: Sheet 1 = VCH (Vision/Clinical), Sheet 2 = GHS (Health Screening). Headers should be in row 2.",
)
st.markdown("</div>", unsafe_allow_html=True)

if uploaded_file is not None:
    try:
        # Check reference file exists
        if not os.path.exists(REFERENCE_FILE):
            st.error(f"❌ Reference file `{REFERENCE_FILE}` not found. Place it in the same directory as this app.")
            st.stop()

        standard_columns = get_standard_columns()

        # Process
        with st.spinner("⏳ Standardising your data..."):
            final_df, logs, mapped_count, blank_count = process_uploaded_file(
                uploaded_file, standard_columns
            )

        # ── Stats ────────────────────────────────────────────────────────
        st.markdown(f"""
        <div class="stat-row">
            <div class="stat-card">
                <p class="stat-number">{len(final_df)}</p>
                <p class="stat-label">Students</p>
            </div>
            <div class="stat-card">
                <p class="stat-number">{mapped_count}</p>
                <p class="stat-label">Columns Mapped</p>
            </div>
            <div class="stat-card">
                <p class="stat-number">{blank_count}</p>
                <p class="stat-label">Columns Blank</p>
            </div>
            <div class="stat-card">
                <p class="stat-number">{len(standard_columns)}</p>
                <p class="stat-label">Total Columns</p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Processing Log ───────────────────────────────────────────────
        st.markdown('<div class="section-header"><h3>📋 Processing Log</h3></div>', unsafe_allow_html=True)

        for level, msg in logs:
            css_class = f"log-{level}"
            icon = {"success": "✓", "info": "ℹ", "warn": "⚠"}.get(level, "•")
            st.markdown(
                f'<div class="log-entry {css_class}">{icon}  {msg}</div>',
                unsafe_allow_html=True,
            )

        # ── Data Preview ─────────────────────────────────────────────────
        st.markdown('<div class="section-header"><h3>👁️ Data Preview</h3></div>', unsafe_allow_html=True)

        # Show only columns that have data
        cols_with_data = [c for c in final_df.columns if final_df[c].replace("", pd.NA).notna().any()]
        preview_df = final_df[cols_with_data]

        st.dataframe(
            preview_df,
            use_container_width=True,
            height=400,
        )

        st.caption(f"Showing {len(cols_with_data)} columns with data (out of {len(standard_columns)} total)")

        # ── Step 2: Decode Encoded Values ─────────────────────────────────
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            border: 1px solid rgba(99, 102, 241, 0.25);
            border-radius: 14px;
            padding: 1.5rem 2rem;
            margin: 2rem 0 1.5rem 0;
        ">
            <h3 style="margin: 0 0 0.5rem 0; font-size: 1.15rem;">🔓 Step 2 — Decode Encoded Values</h3>
            <p style="color: #8b8fa3; margin: 0; font-size: 0.92rem;">
                Some GHS columns contain encoded numbers (e.g. 0 = No, 1 = Yes).
                Enable decoding to replace them with the actual text values.
            </p>
        </div>
        """, unsafe_allow_html=True)

        do_decode = st.checkbox(
            "✅ Decode encoded numeric values to text",
            value=False,
            help="Replaces numeric codes like 0, 1, -98 with their actual text labels (e.g. 'No', 'Yes', 'Not Applicable') for GHS questionnaire columns.",
        )

        if do_decode:
            with st.spinner("🔓 Decoding values..."):
                decoded_df, decode_logs, decoded_count = decode_dataframe(final_df)

            # Decode stats
            st.markdown(f"""
            <div class="stat-row">
                <div class="stat-card">
                    <p class="stat-number">{decoded_count}</p>
                    <p class="stat-label">Columns Decoded</p>
                </div>
                <div class="stat-card">
                    <p class="stat-number">{len(decode)}</p>
                    <p class="stat-label">Total Decode Rules</p>
                </div>
            </div>
            """, unsafe_allow_html=True)

            # Decode log
            with st.expander("📋 Decode Log", expanded=False):
                for level, msg in decode_logs:
                    css_class = f"log-{level}"
                    icon = {"success": "✓", "info": "ℹ", "warn": "⚠"}.get(level, "•")
                    st.markdown(
                        f'<div class="log-entry {css_class}">{icon}  {msg}</div>',
                        unsafe_allow_html=True,
                    )

            # Decoded preview — show only decoded columns
            st.markdown('<div class="section-header"><h3>👁️ Decoded Data Preview</h3></div>', unsafe_allow_html=True)
            std_decode_keys = list(decode.keys())
            decoded_preview_cols = [c for c in std_decode_keys if c in decoded_df.columns]
            if decoded_preview_cols:
                st.dataframe(decoded_df[decoded_preview_cols], use_container_width=True, height=350)
            else:
                st.info("No decoded columns found in the data.")

            # Use decoded version for download
            download_df = decoded_df
            file_suffix = "_standardised_decoded"
        else:
            download_df = final_df
            file_suffix = "_standardised"

        # ── Step 3: Fill Missing Values from Original File ────────────────
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            border: 1px solid rgba(245, 158, 11, 0.25);
            border-radius: 14px;
            padding: 1.5rem 2rem;
            margin: 2rem 0 1.5rem 0;
        ">
            <h3 style="margin: 0 0 0.5rem 0; font-size: 1.15rem;">🔧 Step 3 — Fill Missing Values</h3>
            <p style="color: #8b8fa3; margin: 0; font-size: 0.92rem;">
                If some values are missing in the standardised output, you can fill them
                from the original uploaded file by matching on Child Unique Code.
            </p>
        </div>
        """, unsafe_allow_html=True)

        # Read original sheets for column listing
        uploaded_file.seek(0)
        orig_all_sheets = pd.read_excel(uploaded_file, sheet_name=[0, 1, 2], header=1)
        orig_triage = normalize_multiline_columns(orig_all_sheets[0].copy())
        orig_vch = normalize_multiline_columns(orig_all_sheets[1].copy())
        orig_ghs = normalize_multiline_columns(orig_all_sheets[2].copy())

        sheet_info = {
            "Triage (Sheet 1)": (orig_triage, TRIAGE_MAP),
            "VCH (Sheet 2)": (orig_vch, VCH_MAP),
            "GHS (Sheet 3)": (orig_ghs, GHS_MAP),
        }

        # Session state for accumulated fills, keyed by file name
        fill_key = f"fill_ops_{uploaded_file.name}"
        if fill_key not in st.session_state:
            st.session_state[fill_key] = []

        col_a, col_b = st.columns(2)
        with col_a:
            selected_sheet_name = st.selectbox(
                "📄 Select sheet from original file",
                list(sheet_info.keys()),
                key="fill_sheet_select",
            )

        orig_df, orig_map = sheet_info[selected_sheet_name]
        # Deduplicate column names (keep first) to avoid ambiguous Series errors
        orig_df = orig_df.loc[:, ~orig_df.columns.duplicated(keep='first')]
        id_col = find_id_column(orig_df, orig_map)
        orig_cols = [c for c in orig_df.columns if orig_df[c].notna().any()]

        with col_b:
            ref_col = st.selectbox(
                "📥 Source column (original file)",
                sorted(orig_cols),
                key="fill_ref_col",
            )

        target_col = st.selectbox(
            "📤 Target column (standardised output)",
            sorted(download_df.columns.tolist()),
            key="fill_target_col",
        )

        if id_col is None:
            st.warning("⚠️ Could not find a Child Unique Code / ID column in the selected sheet.")
        else:
            st.caption(f"Matching on: **{id_col}** (original) ↔ **Child Unique Code** (standardised)")

            if st.button("🔄 Fill Missing Values", use_container_width=True):
                st.session_state[fill_key].append({
                    'sheet_name': selected_sheet_name,
                    'id_col': id_col,
                    'ref_col': ref_col,
                    'target_col': target_col,
                })
                st.rerun()

        # Apply all accumulated fills
        for op in st.session_state[fill_key]:
            op_sheet_df, op_map = sheet_info[op['sheet_name']]
            op_id = find_id_column(op_sheet_df, op_map)
            if op_id:
                download_df, _ = fill_missing_from_original(
                    download_df, op_sheet_df, op_id, op['ref_col'], op['target_col']
                )

        # Show fill history
        if st.session_state[fill_key]:
            with st.expander("📋 Fill History", expanded=True):
                for op in st.session_state[fill_key]:
                    st.markdown(
                        f'<div class="log-entry log-success">✓ Filled <b>{op["target_col"]}</b> from <b>{op["ref_col"]}</b> ({op["sheet_name"]})</div>',
                        unsafe_allow_html=True,
                    )
                if st.button("🗑️ Clear all fills", key="clear_fills"):
                    st.session_state[fill_key] = []
                    st.rerun()

            file_suffix += "_filled"

        # ── Download ─────────────────────────────────────────────────────
        st.markdown('<div class="section-header"><h3>⬇️ Download Final Excel</h3></div>', unsafe_allow_html=True)

        file_basename = os.path.splitext(uploaded_file.name)[0]
        output_filename = f"{file_basename}{file_suffix}.xlsx"

        excel_bytes = to_excel_bytes(download_df)

        st.download_button(
            label=f"📥 Download {('Standardised + Decoded' if do_decode else 'Standardised')} Excel",
            data=excel_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.info(f"📁 Output file: **{output_filename}** ({len(download_df)} rows × {len(standard_columns)} columns{' — decoded' if do_decode else ''})")

    except Exception as e:
        st.error(f"❌ Error processing file: {e}")
        import traceback
        st.code(traceback.format_exc())

else:
    # Empty state
    st.markdown("""
    <div style="text-align: center; padding: 3rem 0; color: #6b7280;">
        <p style="font-size: 3rem; margin: 0;">📁</p>
        <p style="font-size: 1.1rem; margin: 0.5rem 0;">Upload a manually-recorded Excel file to get started</p>
        <p style="font-size: 0.85rem; color: #9ca3af;">
            The file should have two sheets — VCH (Vision/Clinical) and GHS (Health Screening)
        </p>
    </div>
    """, unsafe_allow_html=True)
